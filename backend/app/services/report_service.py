from datetime import datetime, time, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from app.models.attendance import Attendance
from app.models.employee import Employee
from app.utils.helpers import get_vn_datetime, format_datetime_vn, get_export_path
from sqlalchemy import and_
from app.models.attendance_recovery import AttendanceRecoveryRequest

# Company configuration rules
COMPANY_CONFIG = {
    'WORK_START_TIME': time(8, 0),      # 8:00 AM
    'WORK_END_TIME': time(17, 0),       # 5:00 PM
    'LUNCH_START': time(12, 0),         # 12:00 PM
    'LUNCH_END': time(13, 0),           # 1:00 PM
    'STANDARD_WORK_HOURS': 8,           # 8 hours/day
    'LATE_THRESHOLD_MINUTES': 15,       # Late > 15 minutes
    'EARLY_THRESHOLD_MINUTES': 15,      # Leave early > 15 minutes
    'HALF_DAY_HOURS': 4,               # 4 hours = half day
    'WORKING_DAYS_PER_WEEK': 5,        # 5 days/week
}

def get_working_days_between(start_date, end_date):
    """Calculates the number of working days (Mon-Fri) within any given period."""
    total = 0
    for i in range((end_date - start_date).days + 1):
        day = start_date + timedelta(days=i)
        if day.weekday() < 5:  # 0-4 are Mon-Fri
            total += 1
    return total

def calculate_work_hours(checkin_time, checkout_time):
    """Calculates work hours, excluding lunch break."""
    if not checkin_time or not checkout_time or checkout_time <= checkin_time:
        return 0

    total_delta = checkout_time - checkin_time
    total_hours = total_delta.total_seconds() / 3600

    lunch_start = datetime.combine(checkin_time.date(), COMPANY_CONFIG['LUNCH_START'])
    lunch_end = datetime.combine(checkin_time.date(), COMPANY_CONFIG['LUNCH_END'])

    # Subtract lunch break if working hours cover the lunch period
    if checkin_time <= lunch_start and checkout_time >= lunch_end:
        total_hours -= (COMPANY_CONFIG['LUNCH_END'].hour - COMPANY_CONFIG['LUNCH_START'].hour) # Subtract lunch hours (e.g., 1 hour)

    return max(0, total_hours)

def calculate_late_minutes(checkin_time):
    """Calculates minutes late."""
    if not checkin_time:
        return 0
    standard_time = datetime.combine(checkin_time.date(), COMPANY_CONFIG['WORK_START_TIME'])
    return max(0, (checkin_time - standard_time).total_seconds() / 60)

def calculate_early_minutes(checkout_time):
    """Calculates minutes left early."""
    if not checkout_time:
        return 0
    standard_time = datetime.combine(checkout_time.date(), COMPANY_CONFIG['WORK_END_TIME'])
    return max(0, (standard_time - checkout_time).total_seconds() / 60)

def calculate_overtime_hours(work_hours):
    """Calculates overtime hours."""
    return max(0, work_hours - COMPANY_CONFIG['STANDARD_WORK_HOURS'])

def calculate_undertime_hours(work_hours):
    """Calculates undertime (missing) hours."""
    return max(0, COMPANY_CONFIG['STANDARD_WORK_HOURS'] - work_hours)

def get_date_range(args):
    """Gets the date range from query parameters or defaults to the current month."""
    start_date_str = args.get('start_date')
    end_date_str = args.get('end_date')
    vn_now = get_vn_datetime()

    default_start = vn_now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calculate the last day of the current month
    if vn_now.month == 12:
        next_month = vn_now.replace(year=vn_now.year + 1, month=1, day=1)
    else:
        next_month = vn_now.replace(month=vn_now.month + 1, day=1)
    default_end = (next_month - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)

    try:
        start = datetime.strptime(start_date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0) if start_date_str else default_start
        end = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999) if end_date_str else default_end

        if start > end:
            return None, None, {"error": "Start date must be before end date"}, 400

    except ValueError:
        return None, None, {"error": "Invalid date format. Use YYYY-MM-DD"}, 400

    return start, end, None, None

def process_employee_attendance_enhanced(employee_id, start_date, end_date):
    
        records = Attendance.query.filter(
            and_(
                Attendance.employee_id == employee_id.upper(),
                Attendance.timestamp >= start_date,
                Attendance.timestamp <= end_date
            )
        ).order_by(Attendance.timestamp.asc()).all()

        # THÊM: Lấy các recovery requests đã được approve
        approved_recoveries = AttendanceRecoveryRequest.query.filter(
            and_(
                AttendanceRecoveryRequest.employee_id == employee_id.upper(),
                AttendanceRecoveryRequest.request_date >= start_date.date(),
                AttendanceRecoveryRequest.request_date <= end_date.date(),
                AttendanceRecoveryRequest.status == 'approved'
            )
        ).all()
        
        approved_recovery_dates = {req.request_date for req in approved_recoveries}

        working_days_in_period = get_working_days_between(start_date, end_date)

        report = {
            "normal_days": 0,
            "late_days": 0,
            "half_days": 0,
            "recovered_days": 0,  # THÊM: Đếm số ngày recovered
            "early_departure_days": 0,
            "forgot_checkout_days": 0,
            "absent_days": 0,
            "total_work_hours": 0,
            "total_overtime_hours": 0,
            "total_undertime_hours": 0,
            "total_late_minutes": 0,
            "total_early_minutes": 0,
            "total_attendance_days": 0,
            "working_days_in_period": working_days_in_period,
            "attendance_rate": 0,
            "avg_checkin_time": None,
            "avg_checkout_time": None,
            "daily_records": []
        }

        daily_grouped_records = defaultdict(list)
        for record in records:
            daily_grouped_records[record.timestamp.date()].append(record)

        checkin_times_for_avg = []
        checkout_times_for_avg = []

        current_day_iterator = start_date.date()
        while current_day_iterator <= end_date.date():
            day_report = {
                "date": current_day_iterator.strftime("%Y-%m-%d"),
                "checkin_time": None,
                "checkout_time": None,
                "work_hours": 0,
                "late_minutes": 0,
                "early_minutes": 0,
                "overtime_hours": 0,
                "undertime_hours": 0,
                "attendance_type": "absent",
                "status": "absent"
            }

            if current_day_iterator.weekday() < 5:  # Only process weekdays (Mon-Fri)
                day_records = daily_grouped_records.get(current_day_iterator)
                
                if day_records:
                    day_records.sort(key=lambda x: x.timestamp)
                    checkin = next((r for r in day_records if r.status == "check-in"), None)
                    checkout = next((r for r in day_records if r.status == "check-out"), None)

                    if checkin:
                        report["total_attendance_days"] += 1
                        day_report["checkin_time"] = checkin.timestamp.strftime("%H:%M:%S")
                        day_report["attendance_type"] = checkin.attendance_type
                        day_report["status"] = "present"

                        # THÊM: Xử lý các loại attendance type khác nhau
                        if checkin.attendance_type == "normal":
                            report["normal_days"] += 1
                        elif checkin.attendance_type == "late":
                            report["late_days"] += 1
                        elif checkin.attendance_type == "half_day":
                            report["half_days"] += 1
                        elif checkin.attendance_type == "recovered":
                            report["recovered_days"] += 1

                        # Tính toán late minutes (trừ khi là recovered)
                        if checkin.attendance_type != "recovered":
                            late_minutes = calculate_late_minutes(checkin.timestamp)
                            day_report["late_minutes"] = round(late_minutes)
                            report["total_late_minutes"] += round(late_minutes)

                        checkin_times_for_avg.append(checkin.timestamp.time())

                        if checkout:
                            day_report["checkout_time"] = checkout.timestamp.strftime("%H:%M:%S")
                            work_hours = calculate_work_hours(checkin.timestamp, checkout.timestamp)
                            day_report["work_hours"] = round(work_hours, 2)
                            report["total_work_hours"] += round(work_hours, 2)

                            # Tính toán early minutes (trừ khi là recovered)
                            if checkout.attendance_type != "recovered":
                                early_minutes = calculate_early_minutes(checkout.timestamp)
                                day_report["early_minutes"] = round(early_minutes)
                                report["total_early_minutes"] += round(early_minutes)

                                if early_minutes > COMPANY_CONFIG['EARLY_THRESHOLD_MINUTES']:
                                    report["early_departure_days"] += 1

                            overtime = calculate_overtime_hours(work_hours)
                            undertime = calculate_undertime_hours(work_hours)
                            day_report["overtime_hours"] = round(overtime, 2)
                            day_report["undertime_hours"] = round(undertime, 2)
                            report["total_overtime_hours"] += round(overtime, 2)
                            report["total_undertime_hours"] += round(undertime, 2)

                            checkout_times_for_avg.append(checkout.timestamp.time())
                        else:
                            report["forgot_checkout_days"] += 1
                            # Nếu recovered và không có checkout, coi như đủ giờ
                            if checkin.attendance_type == "recovered":
                                day_report["work_hours"] = round(COMPANY_CONFIG['STANDARD_WORK_HOURS'], 2)
                                report["total_work_hours"] += round(COMPANY_CONFIG['STANDARD_WORK_HOURS'], 2)
                            else:
                                day_report["undertime_hours"] = round(COMPANY_CONFIG['STANDARD_WORK_HOURS'], 2)
                                report["total_undertime_hours"] += round(COMPANY_CONFIG['STANDARD_WORK_HOURS'], 2)
                else:
                    # THÊM: Check nếu ngày này có recovery request được approve
                    if current_day_iterator in approved_recovery_dates:
                        # Coi như present với type recovered
                        day_report["status"] = "present"
                        day_report["attendance_type"] = "recovered"
                        day_report["work_hours"] = round(COMPANY_CONFIG['STANDARD_WORK_HOURS'], 2)
                        report["recovered_days"] += 1
                        report["total_attendance_days"] += 1
                        report["total_work_hours"] += round(COMPANY_CONFIG['STANDARD_WORK_HOURS'], 2)
                    else:
                        # No records for this workday => absent
                        report["absent_days"] += 1
                        day_report["undertime_hours"] = round(COMPANY_CONFIG['STANDARD_WORK_HOURS'], 2)
                        report["total_undertime_hours"] += round(COMPANY_CONFIG['STANDARD_WORK_HOURS'], 2)
            
            report["daily_records"].append(day_report)
            current_day_iterator += timedelta(days=1)

        # Calculate attendance rate
        report["attendance_rate"] = (report["total_attendance_days"] / working_days_in_period * 100) if working_days_in_period > 0 else 0

        # Calculate average times
        if checkin_times_for_avg:
            avg_checkin_seconds = sum(t.hour * 3600 + t.minute * 60 + t.second for t in checkin_times_for_avg) / len(checkin_times_for_avg)
            report["avg_checkin_time"] = str(timedelta(seconds=int(avg_checkin_seconds)))

        if checkout_times_for_avg:
            avg_checkout_seconds = sum(t.hour * 3600 + t.minute * 60 + t.second for t in checkout_times_for_avg) / len(checkout_times_for_avg)
            report["avg_checkout_time"] = str(timedelta(seconds=int(avg_checkout_seconds)))

        # Round total values
        report["total_work_hours"] = round(report["total_work_hours"], 2)
        report["total_overtime_hours"] = round(report["total_overtime_hours"], 2)
        report["total_undertime_hours"] = round(report["total_undertime_hours"], 2)
        report["attendance_rate"] = round(report["attendance_rate"], 2)

        return report

def get_employee_report_enhanced(employee_id, start_date, end_date):
    """Generates an enhanced employee attendance report."""
    employee = Employee.query.filter_by(employee_id=employee_id.upper()).first()
    if not employee:
        return None, {"error": "Employee not found"}, 404

    if not employee.status:
        return None, {"error": "Employee is inactive"}, 400

    report = process_employee_attendance_enhanced(employee_id, start_date, end_date)

    return {
        "employee": {
            "employee_id": employee.employee_id,
            "full_name": employee.full_name,
            "department": employee.department,
            "position": employee.position
        },
        "report": report,
        "period": {
            "start_date": format_datetime_vn(start_date),
            "end_date": format_datetime_vn(end_date)
        }
    }, None, None

def get_department_report_enhanced(department, start_date, end_date):
    """Generates an enhanced department attendance report."""
    employees = Employee.query.filter_by(department=department, status=True).all()
    if not employees:
        return None, {"error": "No active employees found in this department"}, 404

    total_report = {
        "total_normal_days": 0,
        "total_late_days": 0,
        "total_half_days": 0,
        "total_early_departure_days": 0,
        "total_forgot_checkout_days": 0,
        "total_absent_days": 0,
        "total_work_hours": 0,
        "total_overtime_hours": 0,
        "total_undertime_hours": 0,
        "total_attendance_days": 0,
        "avg_attendance_rate": 0,
        "employee_reports": []
    }

    attendance_rates = []

    for employee in employees:
        employee_report = process_employee_attendance_enhanced(employee.employee_id, start_date, end_date)

        employee_data = {
            "employee_id": employee.employee_id,
            "full_name": employee.full_name,
            "position": employee.position,
            **employee_report
        }

        total_report["employee_reports"].append(employee_data)
        attendance_rates.append(employee_report["attendance_rate"])

        # Accumulate totals
        total_report["total_normal_days"] += employee_report["normal_days"]
        total_report["total_late_days"] += employee_report["late_days"]
        total_report["total_half_days"] += employee_report["half_days"]
        total_report["total_early_departure_days"] += employee_report["early_departure_days"]
        total_report["total_forgot_checkout_days"] += employee_report["forgot_checkout_days"]
        total_report["total_absent_days"] += employee_report["absent_days"]
        total_report["total_work_hours"] += employee_report["total_work_hours"]
        total_report["total_overtime_hours"] += employee_report["total_overtime_hours"]
        total_report["total_undertime_hours"] += employee_report["total_undertime_hours"]
        total_report["total_attendance_days"] += employee_report["total_attendance_days"]

    # Calculate average attendance rate
    total_report["avg_attendance_rate"] = round(sum(attendance_rates) / len(attendance_rates), 2) if attendance_rates else 0

    # Round totals
    total_report["total_work_hours"] = round(total_report["total_work_hours"], 2)
    total_report["total_overtime_hours"] = round(total_report["total_overtime_hours"], 2)
    total_report["total_undertime_hours"] = round(total_report["total_undertime_hours"], 2)

    return {
        "department": department,
        "total_employees": len(employees),
        "report": total_report,
        "period": {
            "start_date": format_datetime_vn(start_date),
            "end_date": format_datetime_vn(end_date)
        }
    }, None, None

def export_to_excel_enhanced(report_type, start_date, end_date):
    """Exports enhanced attendance reports to Excel."""
    try:
        if not report_type or '/' not in report_type:
            return None, {"error": "Invalid report type format. Use 'employee/{id}' or 'department/{name}'"}, 400

        report_category, identifier = report_type.split('/', 1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Helper function to set header style
        def set_header_style(row_cells):
            for cell in row_cells:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

        title_text = f"ATTENDANCE REPORT - {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"

        if report_category == 'employee':
            employee_id = identifier.upper()
            employee = Employee.query.filter_by(employee_id=employee_id).first()
            if not employee:
                return None, {"error": f"Employee {employee_id} not found"}, 404
            if not employee.status:
                return None, {"error": f"Employee {employee_id} is inactive"}, 400

            report_data = process_employee_attendance_enhanced(employee_id, start_date, end_date)

            ws.append([f"{title_text} - {employee.full_name} ({employee.employee_id})"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = center_align

            ws.append([]) # Blank row

            ws.append([f"Department: {employee.department or 'N/A'}", "", f"Position: {employee.position or 'N/A'}"])
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
            ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=12)

            ws.append([]) # Blank row
            ws.append(["ATTENDANCE OVERVIEW"])
            ws['A5'].font = Font(bold=True, size=12)

            summary_data = [
                ["Total Working Days", report_data["working_days_in_period"]],
                ["Attendance Days", report_data["total_attendance_days"]],
                ["Absence Days", report_data["absent_days"]],
                ["Attendance Rate", f"{report_data['attendance_rate']}%"],
                ["Total Working Hours", f"{report_data['total_work_hours']} hours"],
                ["Total Overtime Hours", f"{report_data['total_overtime_hours']} hours"],
                ["Total Missing Hours", f"{report_data['total_undertime_hours']} hours"],
                ["Late days", report_data["late_days"]],
                ["Early Departure Days", report_data["early_departure_days"]],
                ["Forgot checkout days:", report_data["forgot_checkout_days"]]
            ]

            for row_data in summary_data:
                ws.append(row_data)

            ws.append([])
            ws.append([])

            ws.append(["DAILY ATTENDANCE DETAILS"])
            ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
            ws.append([])

            detail_headers = [
                "Date", "Day of Week", "Check-in Time", "Check-out Time", "Working Hours",
                "Late Arrival (mins)", "Early Leave (mins)", "Overtime (hrs)", "Missing Hours (hrs)",
                "Day Type", "Status", "Notes"
            ]
            ws.append(detail_headers)
            set_header_style(ws[ws.max_row])

            weekday_names_en = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

            for daily_record in report_data["daily_records"]:
                current_date = datetime.strptime(daily_record["date"], "%Y-%m-%d").date()
                weekday_name = weekday_names_en[current_date.weekday()]

                notes = []
                if daily_record["status"] == "absent":
                    notes.append("Absent")
                else:
                    if daily_record["late_minutes"] > COMPANY_CONFIG['LATE_THRESHOLD_MINUTES']:
                        notes.append("Late")
                    if daily_record["early_minutes"] > COMPANY_CONFIG['EARLY_THRESHOLD_MINUTES']:
                        notes.append("Early Departure")
                    if not daily_record["checkout_time"]:
                        notes.append("Forgot Check-out")
                    if daily_record["attendance_type"] == "half_day":
                        notes.append("Half Day")
                
                note_text = ", ".join(notes) if notes else "Normal"

                row_data = [
                    current_date.strftime("%d/%m/%Y"),
                    weekday_name,
                    daily_record["checkin_time"] or "N/A",
                    daily_record["checkout_time"] or "N/A",
                    f"{daily_record['work_hours']:.1f}" if daily_record['work_hours'] > 0 else "0",
                    f"{daily_record['late_minutes']:.0f}" if daily_record['late_minutes'] > 0 else "0",
                    f"{daily_record['early_minutes']:.0f}" if daily_record['early_minutes'] > 0 else "0",
                    f"{daily_record['overtime_hours']:.1f}" if daily_record['overtime_hours'] > 0 else "0",
                    f"{daily_record['undertime_hours']:.1f}" if daily_record['undertime_hours'] > 0 else "0",
                    daily_record['attendance_type'].replace("_", " ").title(),
                    daily_record['status'].title(),
                    note_text
                ]
                ws.append(row_data)

            column_widths = [12, 10, 12, 12, 12, 14, 14, 12, 12, 12, 10, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        elif report_category == 'department':
            department_name = identifier
            employees = Employee.query.filter_by(department=department_name, status=True).all()
            if not employees:
                return None, {"error": f"No active employees found in department: {department_name}"}, 404

            ws.append([f"{title_text} - DEPARTMENT: {department_name}"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = center_align

            ws.append([]) # Blank row

            department_headers = [
                "Employee ID", "Full Name", "Position", "Normal Days", "Late Days", "Half Days", "Early Leaves",
                "Forgot Check-outs", "Absent Days", "Total Hours", "Overtime Hours", "Missing Hours",
                "Attendance Rate (%)", "Evaluation"
            ]
            ws.append(department_headers)
            set_header_style(ws[ws.max_row])

            for employee in employees:
                report_data = process_employee_attendance_enhanced(employee.employee_id, start_date, end_date)
                
                evaluation = []
                if report_data["attendance_rate"] < 80:
                    evaluation.append("Poor Attendance")
                if report_data["late_days"] > 5:
                    evaluation.append("Frequently Late")
                if report_data["early_departure_days"] > 3:
                    evaluation.append("Frequently Leaves Early")
                
                evaluation_text = ", ".join(evaluation) if evaluation else "Good"

                row_data = [
                    employee.employee_id,
                    employee.full_name,
                    employee.position or "N/A",
                    report_data["normal_days"],
                    report_data["late_days"],
                    report_data["half_days"],
                    report_data["early_departure_days"],
                    report_data["forgot_checkout_days"],
                    report_data["absent_days"],
                    f"{report_data['total_work_hours']:.1f}",
                    f"{report_data['total_overtime_hours']:.1f}",
                    f"{report_data['total_undertime_hours']:.1f}",
                    f"{report_data['attendance_rate']:.1f}%",
                    evaluation_text
                ]
                ws.append(row_data)

            column_widths = [10, 25, 15, 15, 12, 15, 15, 18, 12, 12, 12, 12, 18, 25]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        else:
            return None, {"error": f"Invalid report category: {report_category}. Use 'employee' or 'department'"}, 400

        export_path = get_export_path()
        os.makedirs(export_path, exist_ok=True)

        timestamp = get_vn_datetime().strftime("%Y%m%d_%H%M%S")
        safe_identifier = identifier.replace('/', '_').replace('\\', '_')
        filename = f"attendance_report_{report_category}_{safe_identifier}_{timestamp}.xlsx"
        filepath = os.path.join(export_path, filename)

        wb.save(filepath)

        if not os.path.exists(filepath) or os.path.getsize(filepath) == 0:
            return None, {"error": "Failed to create or created file is empty"}, 500

        return {
            "message": "Report exported successfully",
            "download_url": f"/api/reports/download/{filename}",
            "filename": filename,
            "file_path": filepath,
            "file_size": os.path.getsize(filepath),
            "report_type": report_category
        }, None, None

    except Exception as e:
        return None, {"error": f"Export failed: {str(e)}"}, 500