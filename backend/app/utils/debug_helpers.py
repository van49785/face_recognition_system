# Debug helper để test export function
# Thêm vào file utils/debug_helpers.py

import os
import logging
from datetime import datetime, timedelta
from app.models.employee import Employee
from app.models.attendance import Attendance
from app.services.report_service import export_to_excel_enhanced, get_date_range
from app.utils.helpers import get_export_path

def debug_export_function():
    """Debug function để kiểm tra export"""
    logger = logging.getLogger(__name__)
    
    print("=== DEBUG EXPORT FUNCTION ===")
    
    # 1. Kiểm tra thư mục export
    export_path = get_export_path()
    print(f"Export path: {export_path}")
    print(f"Export path exists: {os.path.exists(export_path)}")
    
    if not os.path.exists(export_path):
        try:
            os.makedirs(export_path, exist_ok=True)
            print(f"Created export directory: {export_path}")
        except Exception as e:
            print(f"Failed to create export directory: {e}")
            return False
    
    # 2. Kiểm tra permissions
    try:
        test_file = os.path.join(export_path, 'test_permission.txt')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
        print("Write permission: OK")
    except Exception as e:
        print(f"Write permission error: {e}")
        return False
    
    # 3. Kiểm tra data
    employees = Employee.query.filter_by(status=True).all()
    print(f"Total active employees: {len(employees)}")
    
    if not employees:
        print("No employees found!")
        return False
    
    # 4. Test với employee đầu tiên
    test_employee = employees[0]
    print(f"Testing with employee: {test_employee.employee_id} - {test_employee.full_name}")
    
    # 5. Kiểm tra attendance records
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    records = Attendance.query.filter(
        Attendance.employee_id == test_employee.employee_id,
        Attendance.timestamp >= start_date,
        Attendance.timestamp <= end_date
    ).all()
    
    print(f"Attendance records for {test_employee.employee_id}: {len(records)}")
    
    # 6. Test export function
    report_type = f"employee/{test_employee.employee_id}"
    print(f"Testing export with report_type: {report_type}")
    
    try:
        result, error, status = export_to_excel_enhanced(report_type, start_date, end_date)
        
        if error:
            print(f"Export error: {error} (status: {status})")
            return False
        
        print(f"Export result: {result}")
        
        # 7. Kiểm tra file
        if result and 'file_path' in result:
            file_path = result['file_path']
            print(f"Checking file: {file_path}")
            print(f"File exists: {os.path.exists(file_path)}")
            
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                print(f"File size: {file_size} bytes")
                
                if file_size > 0:
                    print("Export test: SUCCESS")
                    return True
                else:
                    print("File is empty!")
                    return False
            else:
                print("File does not exist!")
                return False
        else:
            print("No file path in result!")
            return False
            
    except Exception as e:
        print(f"Export test exception: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_openpyxl():
    """Test openpyxl library"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        print("=== TESTING OPENPYXL ===")
        print(f"OpenPyXL version: {openpyxl.__version__}")
        
        # Test tạo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Test thêm data
        ws.append(["Test Header 1", "Test Header 2"])
        ws.append(["Test Data 1", "Test Data 2"])
        
        # Test formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        
        # Test save
        test_path = get_export_path()
        test_file = os.path.join(test_path, 'test_openpyxl.xlsx')
        
        wb.save(test_file)
        
        if os.path.exists(test_file):
            file_size = os.path.getsize(test_file)
            print(f"Test file created: {test_file} ({file_size} bytes)")
            os.remove(test_file)  # Clean up
            print("OpenPyXL test: SUCCESS")
            return True
        else:
            print("Failed to create test file")
            return False
            
    except Exception as e:
        print(f"OpenPyXL test error: {e}")
        import traceback
        traceback.print_exc()
        return False